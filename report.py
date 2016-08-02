#!/usr/bin/python -u
# coding=utf-8
import sys
reload(sys)
sys.setdefaultencoding("utf8")
import cgitb
cgitb.enable()
import cgi
import json
from datetime import datetime
from dbconn import *
from random import randint
try:
    from collections import OrderedDict
except ImportError:
    from ordereddict import OrderedDict
import util
from openpyxl import load_workbook
form = cgi.FieldStorage()


base = CgBase(util.get_location())
#util.print_header()
#util.print_html_header("Test")
purchases = base.get_purchases()
boxes = base.get_boxes()

box_mapping = {
     1:  4,
     2:  5,
     3:  7,
     5: 11,
     6: 10,
     7:  9,
     8:  8,
    10: 15,
    11: 14,
    12: 13,
    13: 12,
    15: 20,
    16: 19,
    17: 18,
    18: 17,
    20: 24,
    21: 23,
    22: 22,
    23: 21,
    25: 30,
    26: 29,
    27: 28,
    28: 27,
    30: 31,
    31: 32,
    33: 33,
    34: 34,
    36: 35,
    37: 36,
    38: 37,
    40: 38,
    41: 39,
    42: 40,
    44: 41,
    45: 42,
    46: 43,
    48: 44,
    49: 45,
    50: 46,
    52: 49,
    54: 48,
    55: 47,
    56:  2,
    57: 16,
    58:  4,
    59:  5,
    60:  7
}

def colnum_string(n):
    div = n
    string = ""
    temp = 0
    while div > 0:
        module = (div-1) % 26
        string = chr(65+module) + string
        div = int((div-module) / 26)
    return string

def day_to_datetime(day):
    return datetime.strptime(day, '%Y-%m-%d')


br = "<br>"
wb = load_workbook('example.xlsx')
ws = wb.get_sheet_by_name("Foglio1")
#print ws.title, br
#print ws.cell("A2").value, br
#print ws["A2"].value, br

firstday = day_to_datetime("2016-06-20")
stats =  base.get_box_stats()
for date, stat in stats.iteritems():
    diff = (day_to_datetime(date) - firstday).days
#    print diff, date, stat, br
    column = colnum_string(6 + diff * 3)
    for box, quantity in stat.iteritems():
        try:
            row = box_mapping[box]
        except: # for mix which are not to be added in the future
            continue
#        print "---- (", column, ", ", row, ") = ", quantity, br
        ws[column+str(row)] = quantity

wb.save("report.xlsx")

#util.print_html_footer()
print "Location: report.xlsx"
print
